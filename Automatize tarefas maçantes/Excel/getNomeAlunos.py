import openpyxl, os

os.chdir(r'C:\Users\herso\OneDrive\APython\Exercícios\Arquivos')

wb = openpyxl.load_workbook('prazo.xlsx')

#print(wb.get_sheet_names())
 #['Page 1', 'Page 2', 'Page 3']  Retorno do print acima.

sheet1 = wb['Page 1']  # Informação tirada da nova edição do livro!

sheet2 = wb['Page 2']

sheet3 = wb['Page 3']

nomesplanilha = []


for i in range(12,34):
    if 10 <= int(sheet1.cell(row=i ,column=13).value) <= 18:
        print(sheet1.cell(row=i ,column=6).value)

for i in range(3,16):
    if 10 <= int(sheet2.cell(row=i ,column=8).value) <= 18:
        print(sheet2.cell(row=i ,column=4).value)

# Falta saber como pegar esses nomes para colocar para pesquisar no sigaa.

 