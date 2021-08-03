import openpyxl, os

os.chdir(r'C:\Users\herso\OneDrive\APython\Exercícios\Arquivos')

wb = openpyxl.load_workbook('prazo.xlsx')

#print(wb.get_sheet_names())
 #['Page 1', 'Page 2', 'Page 3']  Retorno do print acima.

sheet1 = wb['Page 1']  # Informação tirada da nova edição do livro!

sheet2 = wb['Page 2']



cellTest = sheet1['F28'].value  #Essa chamada eu consigo trazer o valor da célula escolhida.
#print(cellTest)



#Aqui eu atribui essa chamada a uma variável, e consigo depois extrair a linha, coluna e o valor para montar uma frase.
cel = sheet1['F28']
print('Linha {}, Coluna {}, is {}. Ele está na coordenada {}'.format(cel.row, cel.column, cel.value, cel.coordinate))

#Aqui só mais um exemplo que eu posso chamar as linhas e colunas atribuindo um inteiro, sem usar letra/número. Ex. F28 é (row=28 ,column=6).
print(sheet1.cell(row=28 ,column=6).value)

for i in range(12,34):
    print(i, sheet1.cell(row=i ,column=14).value)



# Determinar o tamanho da planilha.
tamanhoLin = sheet1.max_row
tamanhoCol = sheet1.max_column

print(tamanhoLin)



