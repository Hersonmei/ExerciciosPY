'''Código retirado do livro mais atualizado'''

import openpyxl, os


os.chdir(r'C:\Users\herso\OneDrive\APython\Complementos\Resources\automate_online-materials 1e')
wb = openpyxl.load_workbook('produceSales.xlsx')

sheet = wb['Sheet']

#Os tipos de produto e seus preços atualizados.
PRICE_UPDATES = {'Garlic': 3.07, 'Celery': 1.19, 'Lemon': 1.27}

#Todo: Percorre as linhas em um loop e atualiza os preços.
for rowNum in range(2, sheet.max_row): #Pula a primeira linha
    produceName = sheet.cell(rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value == PRICE_UPDATES[produceName]
wb.save('updatedProduceSales.xlsx')

